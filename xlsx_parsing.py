#! /usr/bin/env python

import openpyxl as xls

wb = xls.load_workbook('my_excel.xlsx', use_iterators=True)

all_sheets = wb.get_sheet_names()

count_sets = {}

wb_out = xls.Workbook()

# gather all sheets 
for s in all_sheets:
    sheet_data = wb.get_sheet_by_name(s)
    for row in sheet_data.iter_rows():
        #set the columns to hit based on their name from sheet_data
        data = {
                'Data Set' : row[5].value,
                'Collection Date' : row[7].value
        }

        # gathering only the collection dates which signify the data set has been collected
        collection_dates = data['Collection Date']
        if collection_dates:
            new_collection_date = collection_dates

            # grab the name of the data set that aligns with the collection dates 
            data_set = data['Data Set']
            if data_set:
                new_data_set = data_set

                # combine the two for sanity to make sure it worked
                combo = [new_data_set, new_collection_date]
                # hit the zeroth element then push it up to the count_sets dict and count each time we see a data set 
                # occur in the combo
                counts = combo[0]
                count_sets[counts] = count_sets.get(counts,0) + 1

# create a new list from the key value pairs inside the count_sets dict
hlist = []
for key, val in count_sets.items():
    hlist.append((val, key))

# sort from high to low based on how many of each data set we have collected
hlist.sort(reverse=True)

for key, val in hlist:
    print key, val

# counts the total counts to give an aggregate
vallist = []
for key, val in hlist:
    vallist.append(key)
    total = sum(vallist)

print total, 'Total'
